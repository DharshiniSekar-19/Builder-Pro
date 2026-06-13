from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.template.loader import render_to_string
from projects.models import Project
from activities.models import Activity
from activities.cpm import run_cpm
from activities.pert import PERTAnalyzer
from accounts.decorators import role_required

import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

@login_required
def analytics_dashboard(request):
    projects = Project.objects.all()
    activities = Activity.objects.all()

    total_projects = projects.count()
    total_activities = activities.count()
    completed_activities = activities.filter(status='completed').count()
    critical_activities = activities.filter(is_critical=True).count()

    project_labels = []
    project_budget = []
    project_progress = []
    for p in projects:
        project_labels.append(p.name[:20])
        project_budget.append(float(p.budget))
        project_progress.append(p.progress())

    return render(request, 'analytics/analytics_dashboard.html', {
        'total_projects': total_projects,
        'total_activities': total_activities,
        'completed_activities': completed_activities,
        'critical_activities': critical_activities,
        'project_labels': project_labels,
        'project_budget': project_budget,
        'project_progress': project_progress,
    })

@login_required
def project_report(request, pk):
    project = get_object_or_404(Project, pk=pk)
    activities = Activity.objects.filter(project=project)
    run_cpm(project)
    activities = Activity.objects.filter(project=project)

    return render(request, 'analytics/project_report.html', {
        'project': project,
        'activities': activities,
    })

@login_required
def cpm_report(request, pk):
    project = get_object_or_404(Project, pk=pk)
    critical_path = run_cpm(project)
    activities = Activity.objects.filter(project=project)
    project_duration = sum(a.expected_duration for a in critical_path)

    return render(request, 'analytics/cpm_report.html', {
        'project': project,
        'critical_path': critical_path,
        'non_critical': [a for a in activities if not a.is_critical],
        'project_duration': round(project_duration, 2),
    })

@login_required
def pert_report(request, pk):
    project = get_object_or_404(Project, pk=pk)
    activities = Activity.objects.filter(project=project)

    total_expected = sum(a.expected_duration for a in activities)
    total_variance = sum(a.variance for a in activities)
    total_std = total_variance ** 0.5

    return render(request, 'analytics/pert_report.html', {
        'project': project,
        'activities': activities,
        'total_expected': round(total_expected, 2),
        'total_variance': round(total_variance, 4),
        'total_std': round(total_std, 2),
    })

@login_required
def export_project_excel(request, pk):
    project = get_object_or_404(Project, pk=pk)
    activities = Activity.objects.filter(project=project)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Report"

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Project: {project.name}'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['ID', 'Activity', 'Expected Duration', 'Variance', 'Std Dev', 'Priority', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)

    for row, act in enumerate(activities, 4):
        ws.cell(row=row, column=1, value=act.id)
        ws.cell(row=row, column=2, value=act.name)
        ws.cell(row=row, column=3, value=act.expected_duration)
        ws.cell(row=row, column=4, value=act.variance)
        ws.cell(row=row, column=5, value=act.standard_deviation)
        ws.cell(row=row, column=6, value=act.get_priority_display())
        ws.cell(row=row, column=7, value=act.get_status_display())

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{project.name}_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_activities_excel(request):
    project_id = request.GET.get('project')
    if project_id:
        activities = Activity.objects.filter(project_id=project_id)
    else:
        activities = Activity.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activities"

    headers = ['ID', 'Project', 'Activity', 'Expected', 'Variance', 'Std Dev', 'Priority', 'Status', 'Critical']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, act in enumerate(activities, 2):
        ws.cell(row=row, column=1, value=act.id)
        ws.cell(row=row, column=2, value=act.project.name)
        ws.cell(row=row, column=3, value=act.name)
        ws.cell(row=row, column=4, value=act.expected_duration)
        ws.cell(row=row, column=5, value=act.variance)
        ws.cell(row=row, column=6, value=act.standard_deviation)
        ws.cell(row=row, column=7, value=act.get_priority_display())
        ws.cell(row=row, column=8, value=act.get_status_display())
        ws.cell(row=row, column=9, value='Yes' if act.is_critical else 'No')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="activities_report.xlsx"'
    wb.save(response)
    return response

@login_required
def ajax_chart_data(request):
    projects = Project.objects.all()
    project_labels = [p.name[:20] for p in projects]
    project_progress = [p.progress() for p in projects]
    project_budget = [float(p.budget) for p in projects]
    status_counts = {
        'planning': projects.filter(status='planning').count(),
        'in_progress': projects.filter(status='in_progress').count(),
        'completed': projects.filter(status='completed').count(),
        'delayed': projects.filter(status='delayed').count(),
        'on_hold': projects.filter(status='on_hold').count(),
    }
    return JsonResponse({
        'project_labels': project_labels,
        'project_progress': project_progress,
        'project_budget': project_budget,
        'status_counts': status_counts,
    })
